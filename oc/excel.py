# oc/excel.py
from __future__ import annotations

import re

from typing import Optional, Any
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
import config

SHEET_NAME = "ORDEN DE COMPRA"

# Cabecera
CELL_COTIZACION = "E8"

# Empresa emisora
CELL_EMISOR_NOMBRE = "D3"
CELL_EMISOR_RUT = "D4"
CELL_EMISOR_GIRO = "D5"
CELL_EMISOR_DIRECCION = "D6"
CELL_EMISOR_EXTRA = "D7"

# Proveedor
CELL_PROV_RAZON = "C9"
CELL_PROV_RUT = "C10"
CELL_PROV_GIRO = "C11"
CELL_PROV_DIRECCION = "C12"
CELL_PROV_COMUNA = "C13"
CELL_PROV_CIUDAD = "C14"
CELL_PROV_TELEFONO = "C15"
CELL_PROV_CONTACTO = "C16"

# Pago
CELL_COND_PAGO = "C17"

# Entrega
CELL_ENTREGA_DIRECCION = "C20"
CELL_ENTREGA_CONTACTO = "C21"
CELL_ENTREGA_TELEFONO = "C22"
CELL_ENTREGA_CORREO = "C23"

# Interno
CELL_CC_TEXTO = "B52"
CELL_SOLICITANTE_TEXTO = "B54"

# Items
ITEMS_START_ROW = 25
COL_CODIGO = "B"
COL_CANTIDAD = "C"
COL_DESCRIPCION = "D"
COL_UNITARIO = "E"
COL_TOTAL = "F"

# Totales
CELL_NETO = "F52"
CELL_IVA = "F55"
CELL_TOTAL = "F57"


def _s(x: Any) -> str:
    return ("" if x is None else str(x)).strip()


def _f(x: Any) -> float:
    try:
        if x is None or x == "":
            return 0.0
        return float(x)
    except Exception:
        return 0.0
    
def _safe_filename(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[\\/:*?\"<>|]+", " ", s)   # caracteres prohibidos
    s = re.sub(r"\s+", " ", s).strip()
    return s


def generar_excel_oc(oc, filename_base: str | None = None) -> str:
    plantilla = Path(config.PLANTILLA_OC_XLSX)
    if not plantilla.exists():
        raise FileNotFoundError(f"No se encontró la plantilla: {plantilla}")

    out_dir = Path(getattr(config, "OUTPUTS_DIR", Path("outputs")))
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(plantilla)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    # ... todo tu llenado igual ...

    now = datetime.now()
    if filename_base:
        base = _safe_filename(filename_base)
        name = f"{base}.xlsx"
    else:
        name = f"DRAFT-{now.strftime('%Y%m%d-%H%M%S')}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

    out_path = out_dir / name
    wb.save(out_path)
    return str(out_path)

    # 1) Empresa emisora (viene como dict: oc.empresa)
    empresa = getattr(oc, "empresa", {}) or {}
    ws[CELL_EMISOR_NOMBRE].value = _s(empresa.get("nombre"))
    ws[CELL_EMISOR_RUT].value = _s(empresa.get("rut"))
    ws[CELL_EMISOR_GIRO].value = _s(empresa.get("giro"))
    ws[CELL_EMISOR_DIRECCION].value = _s(empresa.get("direccion"))
    ws[CELL_EMISOR_EXTRA].value = _s(empresa.get("extra"))



    # 2) Proveedor (dict)
    prov = getattr(oc, "proveedor", {}) or {}
    ws[CELL_PROV_RAZON].value = _s(prov.get("razon_social") or prov.get("nombre"))
    ws[CELL_PROV_RUT].value = _s(prov.get("rut"))
    ws[CELL_PROV_GIRO].value = _s(prov.get("giro"))
    ws[CELL_PROV_DIRECCION].value = _s(prov.get("direccion"))
    ws[CELL_PROV_COMUNA].value = _s(prov.get("comuna"))
    ws[CELL_PROV_CIUDAD].value = _s(prov.get("ciudad"))
    ws[CELL_PROV_TELEFONO].value = _s(prov.get("telefono"))
    ws[CELL_PROV_CONTACTO].value = _s(prov.get("persona_contacto") or prov.get("contacto"))

    # 2.5) Cotización
    cot = _s(getattr(oc, "cotizacion", ""))
    ws[CELL_COTIZACION].value = cot if cot else ""

    # 3) Pago
    cond = _s(getattr(oc, "condicion_pago", "")) or _s(getattr(config, "CONDICION_PAGO_DEFAULT", ""))
    ws[CELL_COND_PAGO].value = cond

    # 4) Entrega (dict: oc.entrega)
    entrega = getattr(oc, "entrega", {}) or {}
    ws[CELL_ENTREGA_DIRECCION].value = _s(entrega.get("direccion"))
    ws[CELL_ENTREGA_CONTACTO].value = _s(entrega.get("contacto"))
    ws[CELL_ENTREGA_TELEFONO].value = _s(entrega.get("telefono"))
    ws[CELL_ENTREGA_CORREO].value = _s(entrega.get("correo"))

    # 5) Interno
    cc = _s(getattr(oc, "centro_costo", ""))
    sol = _s(getattr(oc, "solicitante", ""))
    aut = _s(getattr(oc, "autoriza", ""))

    ws[CELL_CC_TEXTO].value = f"CENTRO DE COSTO : {cc}" if cc else ""
    ws[CELL_SOLICITANTE_TEXTO].value = sol if sol else ""

    # 6) Items
    subtotal = 0.0
    items = getattr(oc, "items", []) or []

    for idx, it in enumerate(items):
        row = ITEMS_START_ROW + idx

        if isinstance(it, dict):
            codigo = _s(it.get("codigo"))
            desc = _s(it.get("descripcion") or it.get("nombre") or it.get("detalle"))
            cantidad = _f(it.get("cantidad"))
            unit = _f(it.get("precio_unitario"))
            total = _f(it.get("subtotal")) if it.get("subtotal") is not None else round(cantidad * unit, 2)
        else:
            codigo = _s(getattr(it, "codigo", ""))
            desc = _s(getattr(it, "descripcion", "") or getattr(it, "nombre", "") or getattr(it, "detalle", ""))
            cantidad = _f(getattr(it, "cantidad", 0))
            unit = _f(getattr(it, "precio_unitario", 0))
            total = round(cantidad * unit, 2)

        subtotal += total

        ws[f"{COL_CODIGO}{row}"].value = codigo
        ws[f"{COL_CANTIDAD}{row}"].value = cantidad
        ws[f"{COL_DESCRIPCION}{row}"].value = desc
        ws[f"{COL_UNITARIO}{row}"].value = unit
        ws[f"{COL_TOTAL}{row}"].value = total

    # 7) Totales
    iva = round(subtotal * float(config.IVA_TASA), 2)
    total_pagar = round(subtotal + iva, 2)

    ws[CELL_NETO].value = round(subtotal, 2)
    ws[CELL_IVA].value = iva
    ws[CELL_TOTAL].value = total_pagar

    # 8) Guardar
    now = datetime.now()
    name = f"DRAFT-{now.strftime('%Y%m%d-%H%M%S')}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = out_dir / name
    wb.save(out_path)

    return str(out_path)
